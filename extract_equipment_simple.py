"""
Electrical Equipment Data Extractor - WITH AUTOMATIC COLOR SEQUENCE
Extracts SERVICE SWITCHGEAR and DISTRIBUTION SWITCHGEAR with automatic connection mapping
AND applies equipment colors based on their position in the sequence

COLOR LOGIC (Verified from PDF):
- SERVICE SWITCHGEAR (MVS): Black
- DISTRIBUTION SWITCHGEAR (DSG): Sequential colors based on position
  
  Page 1 Pattern:
  Row 1: Black (DSGAH110 first), Red, Blue, Orange, Pink, Purple, Yellow, [Light Grey not shown]
  Row 2: Red, Blue, Orange, Pink, Purple, Yellow, [Light Grey not shown]
  Row 3: Red, Blue, Orange, Pink, Purple, Yellow, [Light Grey not shown]
  
  Page 2 Pattern:
  Row 1: Red, Blue, Orange, Pink, Purple, Yellow, Light Grey, Black (DSGBH120 last)
  Row 2: Red, Blue, Orange, Pink, Purple, Yellow, [Light Grey not shown]
  Row 3: Red, Blue, Orange, Pink, Purple, Yellow, [Light Grey not shown]
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================================================
# COLOR DEFINITIONS
# ============================================================================

# Define the color sequence for DSG equipment
COLOR_SEQUENCE = [
    ('Black', '333333'),      # 0 - Dark grey/black
    ('Red', 'FFB3B3'),        # 1 - Light red/pink
    ('Blue', 'B3D9FF'),       # 2 - Light blue  
    ('Orange', 'FFD9B3'),     # 3 - Light orange
    ('Pink', 'FFB3E6'),       # 4 - Pink
    ('Purple', 'D9B3FF'),     # 5 - Light purple
    ('Yellow', 'FFFF99'),     # 6 - Light yellow
    ('Light Grey', 'E6E6E6'), # 7 - Light grey
]

# MVS color (Service Switchgear)
MVS_COLOR = '333333'  # Black/Dark grey

# ============================================================================


def get_dsg_color_by_position(page_num, row_num, col_num, total_cols):
    """
    Get DSG color based on position in the diagram
    
    Parameters:
    page_num (int): Page number (0 or 1)
    row_num (int): Row number within page (0, 1, 2)
    col_num (int): Column number (0 to total_cols-1)
    total_cols (int): Total columns in the row
    
    Returns:
    str: Hex color code
    """
    # Page 1, Row 1 (top row): Black first, then sequence
    if page_num == 0 and row_num == 0:
        if col_num == 0:
            return COLOR_SEQUENCE[0][1]  # Black for first position
        else:
            # Shift by 1 for remaining positions
            color_index = col_num % len(COLOR_SEQUENCE)
            return COLOR_SEQUENCE[color_index][1]
    
    # Page 2, Last row, Last column: Black
    elif page_num == 1 and row_num == 0 and col_num == total_cols - 1:
        return COLOR_SEQUENCE[0][1]  # Black for last position
    
    # All other DSG: Follow standard sequence
    else:
        color_index = (col_num + 1) % len(COLOR_SEQUENCE)  # Start from Red (index 1)
        return COLOR_SEQUENCE[color_index][1]


def extract_properties_enhanced(equipment_name, context_text, all_page_text):
    """Enhanced property extraction with multiple strategies"""
    properties = []
    search_text = context_text
    
    if not any(pattern in context_text.upper() for pattern in ['KVA', 'KV', 'A', 'AMP']):
        pattern = re.escape(equipment_name)
        match = re.search(pattern, all_page_text)
        if match:
            start = max(0, match.start() - 300)
            end = min(len(all_page_text), match.end() + 300)
            search_text = all_page_text[start:end]
    
    # Extract KVA ratings
    kva_matches = re.findall(r'\b(\d+)\s*KVA\b', search_text, re.IGNORECASE)
    if kva_matches:
        max_kva = max([int(k) for k in kva_matches])
        properties.append(f"{max_kva}KVA")
    
    # Extract Amperage
    amp_matches = re.findall(r'\b(\d{3,5})\s*(?:A\b|AMP)', search_text, re.IGNORECASE)
    if amp_matches:
        max_amp = max([int(a) for a in amp_matches])
        properties.append(f"{max_amp}A")
    
    # Extract Primary voltage
    primary_volt_matches = re.findall(r'(?:PRIMARY[:\s]+)?(\d+\.?\d*)\s*kV', search_text, re.IGNORECASE)
    if primary_volt_matches:
        properties.append(f"{primary_volt_matches[0]}kV")
    
    # Extract Secondary voltage
    secondary_volt_match = re.search(r'(?:SECONDARY[:\s]+)?([\d]+Y/[\d]+V)', search_text, re.IGNORECASE)
    if secondary_volt_match:
        properties.append(secondary_volt_match.group(1))
    
    if not any('Y/' in p for p in properties):
        voltage_matches = re.findall(r'\b(480|208|240|600)\s*V\b', search_text)
        if voltage_matches:
            properties.append(f"{voltage_matches[0]}V")
    
    return ', '.join(properties) if properties else ''


def extract_with_positions_pdfplumber(pdf_path):
    """Extract equipment with positions and assign colors based on sequence"""
    try:
        import pdfplumber
        
        equipment_data = []
        seen_equipment = set()
        equipment_pattern = r"'([A-Z]{3}[A-Z0-9]{2}\d{3})'"
        
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                print(f"\nProcessing page {page_num + 1}...")
                
                full_page_text = page.extract_text()
                words = page.extract_words(x_tolerance=3, y_tolerance=3)
                
                # Extract equipment and sort by position
                page_equipment = []
                for i, word in enumerate(words):
                    text = word['text']
                    match = re.search(equipment_pattern, text)
                    
                    if match:
                        equipment_name = match.group(1)
                        equipment_type = equipment_name[:3]
                        
                        if equipment_type not in ['MVS', 'DSG']:
                            continue
                        
                        if equipment_name in seen_equipment:
                            continue
                        seen_equipment.add(equipment_name)
                        
                        # Get surrounding text for properties
                        context_start = max(0, i - 10)
                        context_end = min(len(words), i + 40)
                        context_text = ' '.join([w['text'] for w in words[context_start:context_end]])
                        properties = extract_properties_enhanced(equipment_name, context_text, full_page_text)
                        
                        page_equipment.append({
                            'Equipment': equipment_name,
                            'Type': equipment_type,
                            'Properties': properties,
                            'x_position': word['x0'],
                            'y_position': word['top'],
                            'page': page_num
                        })
                
                # Sort by Y position (rows), then X position (columns)
                page_equipment.sort(key=lambda x: (x['y_position'], x['x_position']))
                
                # Assign colors based on position
                # Group by rows (similar Y coordinates)
                rows = []
                current_row = []
                last_y = None
                y_threshold = 50  # pixels tolerance for same row
                
                for eq in page_equipment:
                    if last_y is None or abs(eq['y_position'] - last_y) < y_threshold:
                        current_row.append(eq)
                    else:
                        if current_row:
                            rows.append(current_row)
                        current_row = [eq]
                    last_y = eq['y_position']
                
                if current_row:
                    rows.append(current_row)
                
                # Assign colors to DSG equipment
                for row_idx, row in enumerate(rows):
                    # Skip MVS rows
                    if row[0]['Type'] == 'MVS':
                        for eq in row:
                            eq['Color'] = MVS_COLOR
                            eq['ColorName'] = 'Black'
                            print(f"  ✓ {eq['Equipment']}: Black (MVS)")
                    else:
                        # DSG row - assign sequential colors
                        for col_idx, eq in enumerate(row):
                            color_hex = get_dsg_color_by_position(page_num, row_idx - 1, col_idx, len(row))
                            color_name = next((name for name, hex_code in COLOR_SEQUENCE if hex_code == color_hex), "Unknown")
                            eq['Color'] = color_hex
                            eq['ColorName'] = color_name
                            print(f"  ✓ {eq['Equipment']}: {color_name} (#{color_hex})")
                
                # Add to main list
                equipment_data.extend(page_equipment)
        
        # Sort final list by page, Y, X
        equipment_data.sort(key=lambda x: (x['page'], x['y_position'], x['x_position']))
        
        # Add empty connection fields
        for eq in equipment_data:
            eq['Alternate From'] = ''
            eq['Primary From'] = ''
        
        return equipment_data
        
    except Exception as e:
        print(f"pdfplumber extraction failed: {e}")
        import traceback
        traceback.print_exc()
        return None


def populate_connections(equipment_data):
    """Populate Primary From and Alternate From for DSG equipment"""
    mvs_equipment = [item for item in equipment_data if item['Type'] == 'MVS']
    dsg_equipment = [item for item in equipment_data if item['Type'] == 'DSG']
    
    if not mvs_equipment:
        return equipment_data
    
    # Group DSG by page
    dsg_by_page = {}
    for dsg in dsg_equipment:
        page = dsg['page']
        if page not in dsg_by_page:
            dsg_by_page[page] = []
        dsg_by_page[page].append(dsg)
    
    mvs_equipment.sort(key=lambda x: (x['page'], x['y_position'], x['x_position']))
    
    # Assign connections
    for page, dsgs_on_page in dsg_by_page.items():
        mvs_on_page = [mvs for mvs in mvs_equipment if mvs['page'] == page]
        
        if not mvs_on_page and mvs_equipment:
            mvs_on_page = mvs_equipment[:2] if len(mvs_equipment) >= 2 else mvs_equipment
        
        for dsg in dsgs_on_page:
            if len(mvs_on_page) >= 2:
                dsg['Primary From'] = mvs_on_page[0]['Equipment']
                dsg['Alternate From'] = mvs_on_page[1]['Equipment']
            elif len(mvs_on_page) == 1:
                dsg['Primary From'] = mvs_on_page[0]['Equipment']
    
    return equipment_data


def extract_from_pdf(pdf_path):
    """Extract equipment data from PDF with sequential colors"""
    print("Extracting equipment with sequential color assignment...")
    equipment_data = extract_with_positions_pdfplumber(pdf_path)
    
    if equipment_data is not None and len(equipment_data) > 0:
        print(f"\n✓ Successfully extracted {len(equipment_data)} items")
        colors_found = sum(1 for item in equipment_data if item.get('Color'))
        print(f"✓ Colors assigned to {colors_found} items")
        return equipment_data
    
    return None


def create_excel_file(equipment_data, output_path):
    """Create formatted Excel file WITH COLOR FORMATTING"""
    clean_data = []
    for item in equipment_data:
        clean_data.append({
            'Equipment': item['Equipment'],
            'Type': item['Type'],
            'Properties': item['Properties'],
            'Alternate From': item['Alternate From'],
            'Primary From': item['Primary From'],
            'Color': item.get('Color'),
            'ColorName': item.get('ColorName', '')
        })
    
    df = pd.DataFrame(clean_data)
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Equipment Data'
    
    # Headers
    headers = ['Equipment', 'Type', 'Properties', 'Alternate From', 'Primary From']
    sheet.append(headers)
    
    # Format headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_num in range(1, 6):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data WITH COLOR FORMATTING
    for idx, row_data in enumerate(clean_data):
        row_num = idx + 2
        sheet.append([
            row_data['Equipment'],
            row_data['Type'],
            row_data['Properties'],
            row_data['Alternate From'],
            row_data['Primary From']
        ])
        
        # Apply color if exists
        if row_data.get('Color'):
            color_hex = row_data['Color']
            row_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
            
            for col_num in range(1, 6):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.fill = row_fill
                
                # Adjust text color based on background brightness
                try:
                    r = int(color_hex[0:2], 16)
                    g = int(color_hex[2:4], 16)
                    b = int(color_hex[4:6], 16)
                    brightness = (0.299 * r + 0.587 * g + 0.114 * b)
                    
                    if brightness < 128:
                        cell.font = Font(color='FFFFFF', bold=False)
                    else:
                        cell.font = Font(color='000000', bold=False)
                except Exception as e:
                    print(f"Warning: Invalid color hex '{color_hex}': {e}")
                    cell.font = Font(color='000000', bold=False)
    
    # Column widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    wb.save(output_path)
    
    df_return = df.drop(columns=['Color', 'ColorName'])
    return df_return


def print_summary(df, equipment_data):
    """Print summary with color statistics"""
    print("\n" + "="*70)
    print("EXTRACTION SUMMARY")
    print("="*70)
    
    print(f"\n✓ Total records: {len(df)}")
    
    with_colors = sum(1 for item in equipment_data if item.get('Color'))
    print(f"✓ Equipment with colors: {with_colors}/{len(equipment_data)}")
    
    # Color distribution
    print(f"\nColor Distribution:")
    color_counts = {}
    for item in equipment_data:
        if item.get('ColorName'):
            color_name = item['ColorName']
            if color_name not in color_counts:
                color_counts[color_name] = 0
            color_counts[color_name] += 1
    
    for color_name in sorted(color_counts.keys()):
        count = color_counts[color_name]
        print(f"  {color_name}: {count} items")
    
    print(f"\nEquipment breakdown:")
    
    type_names = {
        'DSG': 'Distribution Switchgear',
        'MVS': 'Service Switchgear'
    }
    
    for eq_type in sorted(df['Type'].unique()):
        count = len(df[df['Type'] == eq_type])
        type_name = type_names.get(eq_type, 'Unknown')
        
        with_props = len([item for item in equipment_data if item['Type'] == eq_type and item['Properties']])
        type_with_colors = len([item for item in equipment_data if item['Type'] == eq_type and item.get('Color')])
        
        if eq_type == 'DSG':
            with_connections = len([item for item in equipment_data 
                                   if item['Type'] == eq_type 
                                   and (item['Primary From'] or item['Alternate From'])])
            print(f"  {eq_type} ({type_name}): {count} items")
            print(f"    - With properties: {with_props}")
            print(f"    - With connections: {with_connections}")
            print(f"    - With colors: {type_with_colors}")
        else:
            print(f"  {eq_type} ({type_name}): {count} items")
            print(f"    - With properties: {with_props}")
            print(f"    - With colors: {type_with_colors}")


def main(pdf_path, output_path):
    """Main extraction function with automatic sequential coloring"""
    print("=" * 70)
    print("Electrical Equipment Data Extractor - SEQUENTIAL COLOR ASSIGNMENT")
    print("Colors assigned automatically based on equipment position")
    print("=" * 70)
    print(f"\nInput PDF: {pdf_path}")
    print(f"Output Excel: {output_path}\n")
    
    print("Color Logic:")
    print("  - MVS (Service Switchgear): Black")
    print("  - DSG (Distribution Switchgear): Sequential colors by position")
    print("  - Page 1, Row 1: Black first, then Red, Blue, Orange, Pink, Purple, Yellow")
    print("  - Page 2, Last position: Black")
    print("  - All other DSG: Red, Blue, Orange, Pink, Purple, Yellow sequence\n")
    
    equipment_data = extract_from_pdf(pdf_path)
    
    if equipment_data is None or len(equipment_data) == 0:
        print("\n✗ No MVS or DSG equipment found")
        return None
    
    equipment_data = populate_connections(equipment_data)
    
    df = create_excel_file(equipment_data, output_path)
    print(f"\n✓ Excel file created with sequential color formatting: {output_path}")
    
    print_summary(df, equipment_data)
    
    print("\n" + "=" * 70)
    print("Extraction completed successfully!")
    print("=" * 70)
    
    return df, equipment_data


if __name__ == "__main__":
    PDF_PATH = 'D:/AI Data housde/Electrical Diagram Extractor/MEDIUM VOLTAGE.pdf'
    OUTPUT_PATH = 'D:/AI Data housde/Electrical Diagram Extractor/equipment_data_sequential_colors.xlsx'
    
    try:
        import pdfplumber
        print("✓ pdfplumber is installed\n")
    except ImportError:
        print("⚠️  pdfplumber not installed!")
        print("Install with: pip install pdfplumber\n")
    
    result = main(PDF_PATH, OUTPUT_PATH)
